from openpyxl import Workbook, load_workbook
import os

ROOM_FILE = "data/rooms.xlsx"
PARTICIPANT_FILE = "data/participants.xlsx"


# ---------------------------
# 최초 실행 시 엑셀 생성
# ---------------------------

def initialize_files():

    if not os.path.exists(ROOM_FILE):

        wb = Workbook()
        ws = wb.active

        ws.append([
            "room_code",
            "created_at"
        ])

        wb.save(ROOM_FILE)

    if not os.path.exists(PARTICIPANT_FILE):

        wb = Workbook()
        ws = wb.active

        ws.append([
            "room_code",
            "name",
            "location",
            "date",
            "start_time",
            "end_time"
        ])

        wb.save(PARTICIPANT_FILE)


# ---------------------------
# 방 생성
# ---------------------------

def save_room(room_code, created_at):

    wb = load_workbook(ROOM_FILE)
    ws = wb.active

    ws.append([
        room_code,
        created_at
    ])

    wb.save(ROOM_FILE)


# ---------------------------
# 참가자 저장
# ---------------------------

def save_participant(
    room_code,
    name,
    location,
    schedules
):

    wb = load_workbook(PARTICIPANT_FILE)
    ws = wb.active

    for schedule in schedules:

        ws.append([
            room_code,
            name,
            location,
            schedule["date"],
            schedule["start"],
            schedule["end"]
        ])

    wb.save(PARTICIPANT_FILE)


# ---------------------------
# 방 참가자 조회
# ---------------------------

def get_room_participants(room_code):

    wb = load_workbook(PARTICIPANT_FILE)
    ws = wb.active

    data = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if row[0] == room_code:

            data.append({
                "room_code": row[0],
                "name": row[1],
                "location": row[2],
                "date": row[3],
                "start": row[4],
                "end": row[5]
            })

    return data
